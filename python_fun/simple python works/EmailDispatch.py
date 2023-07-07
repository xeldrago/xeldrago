import win32com.client as win32
from openpyxl import Workbook
from openpyxl import load_workbook

import os

# Path to the Excel file
excel_file_path = 'E:\\Repos\\vroom-vroom\\python_fun\\simple python works\\file.xlsx'

# Number of attachments per email
num_attachments = 1

# Outlook automation
outlook = win32.Dispatch('Outlook.Application')
namespace = outlook.GetNamespace('MAPI')

# Load the Excel file
workbook = load_workbook(excel_file_path)
sheet = workbook.active

# Get the index of the last column in the sheet
last_column = sheet.max_column

# Add "Sent" column header
sent_column_header = "Sent"
sheet.cell(row=1, column=last_column + 1, value=sent_column_header)

# Iterate over the rows in the Excel sheet
for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    email = row[1]  # To email address
    subject = row[0]  # Subject
    body = row[2]  # Body of the email

    # Dictionary mapping attachment numbers to file names
    attachments = {}
    for j in range(num_attachments):
        attachment_file = row[j+3]
        if attachment_file:
            attachments[j+1] = attachment_file

    mail_item = outlook.CreateItem(0)  # 0 represents mail item

    # Set email properties
    mail_item.Subject = subject
    mail_item.To = email
    mail_item.Body = body

    try:
        # Attach the PDF files
        for attachment_num, attachment_file in attachments.items():
            attachment_path = os.path.join('F:\work related', attachment_file)
            mail_item.Attachments.Add(Source=attachment_path)

        # Send the email
        mail_item.Send()

        # Update the "Sent" column value to "True"
        sheet.cell(row=row_index, column=last_column + 1, value=True)
        print(f"Email sent successfully: {subject} | {email}")
    except Exception as e:
        # Update the "Sent" column value to "False"
        sheet.cell(row=row_index, column=last_column + 1, value=False)
        print(f"Failed to send email: {subject} | {email} - {str(e)}")

# Save the updated workbook
workbook.save(excel_file_path)

print("Emails sent successfully.")
