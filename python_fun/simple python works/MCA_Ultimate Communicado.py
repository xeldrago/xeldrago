import win32com.client as win32
import pywhatkit as kit
from openpyxl import load_workbook
from datetime import datetime
from textblob import TextBlob

def analyze_sentiment(text):
    blob = TextBlob(text)
    polarity = blob.sentiment.polarity
    subjectivity = blob.sentiment.subjectivity

    if polarity > 0 and subjectivity > 0.7:
        return 'Very Positive'
    elif polarity > 0 and subjectivity > 0.5:
        return 'Positive'
    elif polarity > 0 and subjectivity <= 0.5:
        return 'Slightly Positive'
    elif polarity < 0 and subjectivity > 0.7:
        return 'Very Negative'
    elif polarity < 0 and subjectivity > 0.5:
        return 'Negative'
    elif polarity < 0 and subjectivity <= 0.5:
        return 'Slightly Negative'
    elif subjectivity < 0.5:
        return 'Neutral'
    elif polarity > 0 and subjectivity > 0.3:
        return 'Happy'
    elif polarity < 0 and subjectivity > 0.3:
        return 'Sad'
    elif polarity < 0 and subjectivity > 0.7:
        return 'Angry'
    else:
        return 'Mixed'
    
# Load the Excel workbook
workbook = load_workbook('E:\\Repos\\vroom-vroom\\python_fun\\simple python works\\data.xlsx')
sheet = workbook.active

# Outlook automation
outlook = win32.Dispatch('Outlook.Application')
namespace = outlook.GetNamespace('MAPI')

# Get the current date and time
today = datetime.now()
sentiment_column = sheet.max_column + 1
# Iterate over the rows in the Excel sheet
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
    name, email, phone_number, subject, body, birth_date = row
    if isinstance(birth_date, str):
        birth_date = datetime.strptime(birth_date, '%d.%m.%Y')
    

    # Sending email
    mail_item = outlook.CreateItem(0)
    mail_item.Subject = subject
    mail_item.To = email
    mail_item.Body = body

    try:
        # Send the email
        mail_item.Send()
        print(f"Email sent successfully to {name} at {email}!")

    except Exception as e:
        print(f"Failed to send email to {name} at {email} - {str(e)}")

    
    # Sending Whatsapp message

    if birth_date.month == today.month and birth_date.day == today.day:
        message = f"{subject}\n{body}"
        try:
            # Remove special characters from phone number
            phone_number = ''.join(filter(str.isdigit, str(phone_number)))
            
            country_code = "+91"  # Country code for India
            phone_number_with_country_code = country_code + phone_number
            
            kit.sendwhatmsg_instantly(phone_number_with_country_code, message)
            print(f"Sent birthday message to {name} at {phone_number_with_country_code}!")
        except Exception as e:
            print(f"Failed to send Whatsapp message to {name} at {phone_number_with_country_code} - {str(e)}")


    subject_sentiment = analyze_sentiment(subject)
    body_sentiment = analyze_sentiment(body)

    # Assign sentiment labels to Excel sheet
    sentiment_labels = f"Subject: {subject_sentiment}, Body: {body_sentiment}"

    sheet.cell(row=row_idx + 1, column=sentiment_column, value=sentiment_labels)

# Save the modified workbook
workbook.save('E:\\Repos\\vroom-vroom\\python_fun\\simple python works\\data.xlsx')
