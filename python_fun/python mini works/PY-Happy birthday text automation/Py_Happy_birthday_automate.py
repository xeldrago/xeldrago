import openpyxl
from datetime import datetime, timedelta
import pywhatkit as kit

def set_reminder(description, reminder_date, phone_number):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook('reminders.xlsx')
    sheet = workbook.active

    # Add a new row with the reminder details
    sheet.append([description, reminder_date, phone_number])

    # Save the updated workbook
    workbook.save('reminders.xlsx')

def send_whatsapp_message(message, phone_number, delay_minutes=1):
    # Get the current date and time
    today = datetime.now()

    # Send the WhatsApp message
    kit.sendwhatmsg(phone_number, message, today.hour, today.minute + delay_minutes)

# Example: Set a reminder and send a message
#set_reminder("Dunzo - yes i made it in python ...sad", datetime(2025, 1, 18, 12, 22), "+919994198795")

# Load the Excel workbook
workbook = openpyxl.load_workbook('reminders.xlsx')
sheet = workbook.active

# Get the current date and time
today = datetime.now()

# Loop through the rows in the Excel sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    description, reminder_date, phone_number = row

    # Handle missing or invalid reminder_date
    if reminder_date is None:
        print(f"Skipping row with missing reminder date for '{description}'")
        continue  # Skip this row if the reminder_date is None
    try:
        # Ensure reminder_date is a datetime object
        if isinstance(reminder_date, datetime):
            reminder_date = reminder_date
        else:
            reminder_date = datetime.strptime(str(reminder_date), "%Y-%m-%d %H:%M:%S")  # Convert string to datetime

        reminder_date = datetime.combine(reminder_date, datetime.min.time())  # Combine with a time for comparison
    except Exception as e:
        print(f"Invalid date format for '{description}'. Error: {e}")
        continue  # Skip this row if reminder_date is not valid

    if today.date() == reminder_date.date():
        # If today is the reminder date, send a message to the specified WhatsApp number
        message = f"Reminder: {description}"
        send_whatsapp_message(message, phone_number)
        print(f"Sent reminder for '{description}' to {phone_number}!")
