import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

# Function to save data to Excel
def save_data():
    data = []
    for entry in entries[:-1]:  # Exclude the Remarks entry
        data.append(entry.get())

    row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S")] + data

    ws.append(row)
    wb.save("data.xlsx")
    messagebox.showinfo("Information", "Data saved successfully")

# Function to edit previously entered data
def edit_data(row_num):
    # Load data for selected row into entry fields
    for i, entry in enumerate(entries[:-1]):  # Exclude the Remarks entry
        entry.delete(0, tk.END)
        entry.insert(0, ws.cell(row=row_num, column=i+1).value)

    # Remove old row from Excel file
    ws.delete_rows(row_num)
    wb.save("data.xlsx")
    messagebox.showinfo("Information", "Data edited successfully")

# Function to add new entry fields for previous commissionerate 1
def add_prev_commissionerate_fields():
    prev_commissionerate_fields = [
        tk.Label(app, text="Date of joining in previous commissionerate 1"),
        tk.Entry(app),
        tk.Label(app, text="Date of relief in the previous commissionerate 1"),
        tk.Entry(app)
    ]
    prev_commissionerate_fields[0].grid(row=len(column_names)+2, column=0)
    prev_commissionerate_fields[1].grid(row=len(column_names)+2, column=1)
    prev_commissionerate_fields[2].grid(row=len(column_names)+3, column=0)
    prev_commissionerate_fields[3].grid(row=len(column_names)+3, column=1)

# Create the Excel workbook and worksheet
wb = Workbook()
ws = wb.active
ws.append(["Timestamp", "Employee code", "Name", "DOB", "designation", "present commissionerate",
           "date of joining in present commissionerate in present grade", "date of first joining",
           "date of joining in the present commissionerate irrespective of grade",
           "date of joining in previous commissionerate 1", "date of relief in the previous commissionerate 1",
           "date of joining in previous commissionerate 2", "date of relief in the previous commissionerate 2",
           "date of joining in previous commissionerate 3", "date of relief in the previous commissionerate 3",
           "date of joining in previous commissionerate 4", "date of relief in the previous commissionerate 4",
           "Remarks"])

# Load existing data from the Excel file
try:
    wb_existing = load_workbook("data.xlsx")
    ws_existing = wb_existing.active

    for row in ws_existing.iter_rows(values_only=True):
        ws.append(row)
except FileNotFoundError:
    pass

# Create the tkinter application
app = tk.Tk()
app.title("Data Entry")

column_names = ["Employee code", "Name", "DOB", "designation", "present commissionerate",
                "date of joining in present commissionerate in present grade", "date of first joining",
                "date of joining in the present commissionerate irrespective of grade",
                "date of joining in previous commissionerate 1", "date of relief in the previous commissionerate 1",
                "date of joining in previous commissionerate 2", "date of relief in the previous commissionerate 2",
                "date of joining in previous commissionerate 3", "date of relief in the previous commissionerate 3",
                "date of joining in previous commissionerate 4", "date of relief in the previous commissionerate 4",
                "Remarks"]

entries = []
# Validation functions for employee code and date fields
def validate_employee_code(new_value):
    if len(new_value) > 10:
        return False
    return True

def validate_date(new_value):
    if new_value in ['', '-']:  # Allow empty string and dash input
        return True
    try:
        datetime.strptime(new_value, "%Y-%m-%d")
        return True
    except ValueError:
        return False

# Function to validate the DOB entry
def validate_dob_input(event):
    new_value = event.widget.get()
    if new_value in ['', '-']:
        return True
    try:
        datetime.strptime(new_value, "%Y-%m-%d")
        return True
    except ValueError:
        return False

# ...

for i, column_name in enumerate(column_names):
    label = tk.Label(app, text=column_name)
    label.grid(row=i, column=0)
    entry = tk.Entry(app)
    entry.grid(row=i, column=1)
    entries.append(entry)

    # Add validation to certain fields
    if column_name == 'Employee code':
        vcmd = app.register(validate_employee_code)
        entry.config(validate="key", validatecommand=(vcmd, '%P'))
    
    if column_name == 'DOB':
        entry.bind('<KeyRelease>', validate_dob_input)


# Add Remarks label and entry field
remarks_label = tk.Label(app, text="Extra Remarks")
remarks_label.grid(row=len(column_names)+1, column=0)
remarks_entry = tk.Entry(app)
remarks_entry.grid(row=len(column_names)+1, column=1)
entries.append(remarks_entry)

add_prev_commissionerate_button = tk.Button(app, text="Add Previous Commissionerate Fields",
                                            command=add_prev_commissionerate_fields)
add_prev_commissionerate_button.grid(row=len(column_names)+2, column=1)

save_button = tk.Button(app, text="Save", command=save_data)
save_button.grid(row=len(column_names)+3, column=1)

# Bind double-click event to the rows in the Excel view
def on_double_click(event):
    selected_row = tree.selection()[0]
    row_number = int(selected_row.lstrip("I"))
    edit_data(row_number)

app.mainloop()
